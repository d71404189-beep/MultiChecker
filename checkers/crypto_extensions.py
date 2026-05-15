# -*- coding: utf-8 -*-
"""
Расширения для CryptoChecker v1.0.52
- Проверка всех ERC-20 токенов
- Проверка NFT
- Экспорт в Excel
"""

import aiohttp
import time
from typing import List, Dict, Any

# ═══════════════════════════════════════════════════════════════════════════
#  ПРОВЕРКА ВСЕХ ERC-20 ТОКЕНОВ
# ═══════════════════════════════════════════════════════════════════════════

async def get_all_erc20_tokens(address: str, chain: str, session: aiohttp.ClientSession, timeout: int = 10) -> List[Dict]:
    """
    Получить ВСЕ ERC-20 токены на адресе через Etherscan API
    
    Args:
        address: Ethereum адрес
        chain: ethereum, bsc, polygon
        session: aiohttp сессия
        timeout: таймаут запроса
    
    Returns:
        List[Dict]: Список токенов с балансами
    """
    
    # API endpoints для разных сетей
    api_urls = {
        "ethereum": "https://api.etherscan.io/api",
        "bsc": "https://api.bscscan.com/api",
        "polygon": "https://api.polygonscan.com/api",
    }
    
    api_url = api_urls.get(chain)
    if not api_url:
        return []
    
    # Получаем список всех токенов
    url = f"{api_url}?module=account&action=tokentx&address={address}&page=1&offset=100&sort=desc"
    
    try:
        async with session.get(url, timeout=timeout) as resp:
            if resp.status != 200:
                return []
            
            data = await resp.json()
            if data.get("status") != "1":
                return []
            
            transactions = data.get("result", [])
            if not transactions:
                return []
            
            # Собираем уникальные контракты токенов
            token_contracts = {}
            for tx in transactions:
                contract = tx.get("contractAddress", "").lower()
                if contract and contract not in token_contracts:
                    token_contracts[contract] = {
                        "symbol": tx.get("tokenSymbol", "UNKNOWN"),
                        "name": tx.get("tokenName", "Unknown Token"),
                        "decimals": int(tx.get("tokenDecimal", 18)),
                    }
            
            # Проверяем баланс каждого токена
            tokens_with_balance = []
            for contract, token_info in token_contracts.items():
                balance = await get_token_balance(address, contract, token_info["decimals"], api_url, session, timeout)
                if balance > 0:
                    # Получаем цену токена
                    price_usd = await get_token_price(token_info["symbol"], session, timeout)
                    value_usd = balance * price_usd if price_usd else 0
                    
                    tokens_with_balance.append({
                        "contract": contract,
                        "symbol": token_info["symbol"],
                        "name": token_info["name"],
                        "balance": balance,
                        "decimals": token_info["decimals"],
                        "price_usd": price_usd,
                        "value_usd": value_usd,
                    })
            
            # Сортируем по стоимости
            tokens_with_balance.sort(key=lambda x: x["value_usd"], reverse=True)
            
            return tokens_with_balance
            
    except Exception as e:
        print(f"Ошибка получения токенов: {e}")
        return []


async def get_token_balance(address: str, contract: str, decimals: int, api_url: str, session: aiohttp.ClientSession, timeout: int) -> float:
    """Получить баланс конкретного токена"""
    url = f"{api_url}?module=account&action=tokenbalance&contractaddress={contract}&address={address}&tag=latest"
    
    try:
        async with session.get(url, timeout=timeout) as resp:
            if resp.status != 200:
                return 0
            
            data = await resp.json()
            if data.get("status") != "1":
                return 0
            
            balance_wei = int(data.get("result", 0))
            balance = balance_wei / (10 ** decimals)
            
            return balance
    except Exception:
        return 0


async def get_token_price(symbol: str, session: aiohttp.ClientSession, timeout: int) -> float:
    """Получить цену токена через CoinGecko"""
    # Кэш цен
    if not hasattr(get_token_price, "cache"):
        get_token_price.cache = {}
        get_token_price.cache_time = {}
    
    # Проверяем кэш (TTL 5 минут)
    if symbol in get_token_price.cache:
        if time.time() - get_token_price.cache_time.get(symbol, 0) < 300:
            return get_token_price.cache[symbol]
    
    # Маппинг популярных токенов
    symbol_to_id = {
        "USDT": "tether",
        "USDC": "usd-coin",
        "DAI": "dai",
        "WETH": "weth",
        "WBTC": "wrapped-bitcoin",
        "LINK": "chainlink",
        "UNI": "uniswap",
        "AAVE": "aave",
        "SHIB": "shiba-inu",
        "PEPE": "pepe",
        "MATIC": "matic-network",
        "BNB": "binancecoin",
    }
    
    coin_id = symbol_to_id.get(symbol.upper())
    if not coin_id:
        # Пробуем поиск по символу
        try:
            search_url = f"https://api.coingecko.com/api/v3/search?query={symbol}"
            async with session.get(search_url, timeout=timeout) as resp:
                if resp.status == 200:
                    data = await resp.json()
                    coins = data.get("coins", [])
                    if coins:
                        coin_id = coins[0].get("id")
        except Exception:
            pass
    
    if not coin_id:
        return 0
    
    # Получаем цену
    try:
        price_url = f"https://api.coingecko.com/api/v3/simple/price?ids={coin_id}&vs_currencies=usd"
        async with session.get(price_url, timeout=timeout) as resp:
            if resp.status == 200:
                data = await resp.json()
                price = data.get(coin_id, {}).get("usd", 0)
                
                # Сохраняем в кэш
                get_token_price.cache[symbol] = price
                get_token_price.cache_time[symbol] = time.time()
                
                return price
    except Exception:
        pass
    
    return 0


# ═══════════════════════════════════════════════════════════════════════════
#  ПРОВЕРКА NFT
# ═══════════════════════════════════════════════════════════════════════════

async def get_nfts(address: str, chain: str, session: aiohttp.ClientSession, timeout: int = 10) -> List[Dict]:
    """
    Получить NFT на адресе через OpenSea API
    
    Args:
        address: Ethereum адрес
        chain: ethereum, polygon, bsc
        session: aiohttp сессия
        timeout: таймаут запроса
    
    Returns:
        List[Dict]: Список NFT с информацией
    """
    
    # OpenSea API v2
    chain_map = {
        "ethereum": "ethereum",
        "polygon": "matic",
        "bsc": "bsc",
    }
    
    opensea_chain = chain_map.get(chain, "ethereum")
    url = f"https://api.opensea.io/api/v2/chain/{opensea_chain}/account/{address}/nfts"
    
    headers = {
        "Accept": "application/json",
        "X-API-KEY": ""  # Можно работать без ключа, но с лимитами
    }
    
    try:
        async with session.get(url, headers=headers, timeout=timeout) as resp:
            if resp.status != 200:
                return []
            
            data = await resp.json()
            nfts_data = data.get("nfts", [])
            
            if not nfts_data:
                return []
            
            nfts = []
            for nft in nfts_data:
                collection = nft.get("collection", "Unknown")
                name = nft.get("name", "Unnamed")
                token_id = nft.get("identifier", "")
                image_url = nft.get("image_url", "")
                
                # Получаем floor price коллекции
                floor_price = await get_collection_floor_price(collection, opensea_chain, session, timeout)
                
                nfts.append({
                    "collection": collection,
                    "name": name,
                    "token_id": token_id,
                    "image_url": image_url,
                    "floor_price_eth": floor_price,
                    "floor_price_usd": floor_price * await get_eth_price(session, timeout) if floor_price else 0,
                })
            
            return nfts
            
    except Exception as e:
        print(f"Ошибка получения NFT: {e}")
        return []


async def get_collection_floor_price(collection_slug: str, chain: str, session: aiohttp.ClientSession, timeout: int) -> float:
    """Получить floor price коллекции"""
    # Кэш floor prices
    if not hasattr(get_collection_floor_price, "cache"):
        get_collection_floor_price.cache = {}
        get_collection_floor_price.cache_time = {}
    
    cache_key = f"{collection_slug}_{chain}"
    
    # Проверяем кэш (TTL 10 минут)
    if cache_key in get_collection_floor_price.cache:
        if time.time() - get_collection_floor_price.cache_time.get(cache_key, 0) < 600:
            return get_collection_floor_price.cache[cache_key]
    
    url = f"https://api.opensea.io/api/v2/collections/{collection_slug}/stats"
    
    headers = {
        "Accept": "application/json",
    }
    
    try:
        async with session.get(url, headers=headers, timeout=timeout) as resp:
            if resp.status != 200:
                return 0
            
            data = await resp.json()
            floor_price = data.get("total", {}).get("floor_price", 0)
            
            # Сохраняем в кэш
            get_collection_floor_price.cache[cache_key] = floor_price
            get_collection_floor_price.cache_time[cache_key] = time.time()
            
            return floor_price
    except Exception:
        return 0


async def get_eth_price(session: aiohttp.ClientSession, timeout: int) -> float:
    """Получить текущую цену ETH"""
    if not hasattr(get_eth_price, "cache"):
        get_eth_price.cache = 0
        get_eth_price.cache_time = 0
    
    # Проверяем кэш (TTL 1 минута)
    if time.time() - get_eth_price.cache_time < 60:
        return get_eth_price.cache
    
    try:
        url = "https://api.coingecko.com/api/v3/simple/price?ids=ethereum&vs_currencies=usd"
        async with session.get(url, timeout=timeout) as resp:
            if resp.status == 200:
                data = await resp.json()
                price = data.get("ethereum", {}).get("usd", 0)
                
                get_eth_price.cache = price
                get_eth_price.cache_time = time.time()
                
                return price
    except Exception:
        pass
    
    return 0


# ═══════════════════════════════════════════════════════════════════════════
#  ЭКСПОРТ В EXCEL
# ═══════════════════════════════════════════════════════════════════════════

def export_to_excel(results: List[Dict], filename: str = "multichecker_results.xlsx"):
    """
    Экспортировать результаты в Excel с форматированием
    
    Args:
        results: Список результатов проверки
        filename: Имя файла для сохранения
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        
        # Создаем workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Results"
        
        # Заголовки
        headers = ["#", "Type", "Input", "Status", "Balance USD", "Details", "Timestamp"]
        ws.append(headers)
        
        # Форматирование заголовков
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Данные
        for idx, result in enumerate(results, 1):
            row_data = [
                idx,
                result.get("type", "unknown"),
                result.get("input", "")[:50],
                "✓ Valid" if result.get("exists") else "✗ Invalid",
                result.get("info", {}).get("total_usd", 0),
                result.get("info", {}).get("message", "")[:100],
                result.get("timestamp", ""),
            ]
            ws.append(row_data)
            
            # Цветовое кодирование
            status_cell = ws.cell(row=idx+1, column=4)
            if result.get("exists"):
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100", bold=True)
            else:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006")
        
        # Автоширина колонок
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Сохраняем
        wb.save(filename)
        return f"✓ Экспортировано в {filename}"
        
    except Exception as e:
        return f"✗ Ошибка экспорта: {e}"
